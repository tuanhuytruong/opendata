"""Read-only dataset profiling and chart API for OpenData report runs."""
from __future__ import annotations

import base64
import binascii
import csv
import hashlib
import hmac
import html
import io
import json
import math
import os
from collections import Counter
import urllib.error
import urllib.request
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal, cast
from uuid import uuid4

import duckdb
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError

from database_adapters import read_registered_source
from formatting import compact_number, format_display_date, format_number, parse_date_value, percent, value_format_descriptor
from planning import analyst_proposals, business_semantic_catalog, canonical_field_name, display_label, evidence_for_chart, is_starter_analysis_request, narrative_from_evidence, parse_filter, propose_charts
from source_registry import public_source, registered_sources
from run_store import DurableJobQueue, RunStore

MAX_UPLOAD_BYTES = 100 * 1024 * 1024
MAX_PROFILE_ROWS = 600_000
MAX_REQUESTS_PER_MINUTE = 240
MAX_EDA_COLUMNS = 100
MAX_EDA_TOP_CATEGORIES = 10
MAX_DATA_PAGE_SIZE = 100
MAX_DATA_EXPORT_ROWS = 10_000
MAX_DATA_SEARCH_LENGTH = 200
# Attach profiles inspect a bounded, deterministic sample. Full profiles are computed
# only when a caller explicitly asks for one of the richer analysis endpoints.
ATTACH_PROFILE_SAMPLE_ROWS = 1_000
DATA_DIR = Path(__file__).resolve().parents[2] / "var" / "uploads"
JOB_DIR = Path(__file__).resolve().parents[2] / "var" / "jobs"
STATIC_DIR = Path(os.getenv("OPENDATA_STATIC_DIR", str(Path(__file__).resolve().parents[2] / "dist")))
RUN_STORE = RunStore(DATA_DIR)
JOB_QUEUE = DurableJobQueue(JOB_DIR)
PROFILE_CACHE: dict[str, DatasetProfile] = {}
VALID_CHARTS = {"bar", "line", "area", "scatter"}


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "no-referrer"
        response.headers["Cache-Control"] = "no-store"
        return response


class BasicAuthMiddleware(BaseHTTPMiddleware):
    """Optional pilot-only HTTP Basic Auth, configured exclusively by environment."""

    async def dispatch(self, request: Request, call_next):
        username = os.getenv("OPENDATA_BASIC_AUTH_USER", "").strip()
        password = os.getenv("OPENDATA_BASIC_AUTH_PASSWORD", "")
        if not username and not password:
            return await call_next(request)
        if not username or not password:
            return JSONResponse({"detail": "Basic Auth configuration is invalid."}, status_code=503)
        authorization = request.headers.get("authorization", "")
        valid = False
        if authorization.startswith("Basic "):
            try:
                decoded = base64.b64decode(authorization[6:], validate=True).decode("utf-8")
                supplied_user, supplied_password = decoded.split(":", 1)
                valid = hmac.compare_digest(supplied_user, username) and hmac.compare_digest(supplied_password, password)
            except (ValueError, UnicodeDecodeError, binascii.Error):
                valid = False
        if not valid:
            return JSONResponse(
                {"detail": "Authentication required."},
                status_code=401,
                headers={"WWW-Authenticate": 'Basic realm="OpenData pilot", charset="UTF-8"'},
            )
        return await call_next(request)


class InMemoryRateLimitMiddleware(BaseHTTPMiddleware):
    """Conservative pilot safeguard; replace with shared rate limiting before scale-out."""
    buckets: dict[str, list[float]] = {}

    async def dispatch(self, request: Request, call_next):
        if request.url.path in {"/api/health", "/api/readiness"}:
            return await call_next(request)
        now = datetime.now(timezone.utc).timestamp()
        client = request.client.host if request.client else "unknown"
        active = [stamp for stamp in self.buckets.get(client, []) if stamp > now - 60]
        if len(active) >= MAX_REQUESTS_PER_MINUTE:
            return JSONResponse({"detail": "Request rate limit exceeded. Try again shortly."}, status_code=429)
        self.buckets[client] = [*active, now]
        return await call_next(request)


app = FastAPI(title="OpenData Report API", version="0.3.0")
app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(InMemoryRateLimitMiddleware)
app.add_middleware(BasicAuthMiddleware)
app.add_middleware(CORSMiddleware, allow_origins=["http://localhost:5173", "http://localhost:5174"], allow_credentials=False, allow_methods=["GET", "POST", "DELETE"], allow_headers=["content-type"])


class ColumnProfile(BaseModel):
    name: str
    kind: Literal["time", "num", "cat", "id", "unknown"]
    null_count: int
    null_ratio: float
    distinct_count: int
    description: str


class DatasetProfile(BaseModel):
    run_id: str
    file_name: str
    row_count: int
    column_count: int
    usable_column_count: int
    columns: list[ColumnProfile]
    warnings: list[str]
    preview: list[dict[str, str]]
    # "sampled" is safe to render immediately after attach; "complete" means the
    # retained dataset has received the full, cached profile.
    profile_status: Literal["sampled", "complete"] = "complete"
    profiled_row_count: int = 0


class FilterSpec(BaseModel):
    column: str
    operator: Literal["equals", "not_equals", "greater_than", "greater_or_equal", "less_than", "less_or_equal"] = "equals"
    value: str = Field(min_length=1, max_length=500)


class DataQuery(BaseModel):
    """Validated raw-data query; filters are schema fields, never client SQL."""
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=25, ge=10, le=MAX_DATA_PAGE_SIZE)
    search: str = Field(default="", max_length=MAX_DATA_SEARCH_LENGTH)
    sort_by: str | None = Field(default=None, max_length=160)
    sort_direction: Literal["asc", "desc"] = "asc"
    filters: list[FilterSpec] = Field(default_factory=list, max_length=10)


class ChartRequest(BaseModel):
    dimension: str
    metric: str
    aggregation: Literal["sum", "avg", "count"] = "sum"
    chart_type: Literal["bar", "line", "area", "pie", "donut", "scatter", "pareto", "stacked_bar", "heatmap"] = "bar"
    # Scatter is a genuine two-measure aggregate: dimension supplies safe labels,
    # metric is Y and x_metric is X. Other charts ignore x_metric.
    x_metric: str | None = None
    secondary_dimension: str | None = None
    # When grouping by a secondary dimension, retain the ranked limit inside each
    # group rather than applying one global limit to all groups.
    limit_per_secondary: bool = False
    limit: int = Field(default=12, ge=1, le=30)
    filters: list[FilterSpec] = Field(default_factory=list, max_length=10)


class ReportRequest(BaseModel):
    """Deprecated compatibility input; converted into the persisted run document."""
    title: str = Field(default="OpenData Analytics Report", min_length=1, max_length=120)
    charts: list[ChartRequest] = Field(min_length=1, max_length=12)


class ReportSection(BaseModel):
    section_id: str = Field(min_length=1, max_length=80)
    heading: str = Field(default="Untitled section", min_length=1, max_length=160)
    commentary: str = Field(default="", max_length=8_000)
    recommended_actions: list[str] = Field(default_factory=list, max_length=20)


class ManualGlossaryNote(BaseModel):
    note_id: str = Field(min_length=1, max_length=80)
    text: str = Field(min_length=1, max_length=1_000)


class CustomReportArtifact(BaseModel):
    artifact_id: str = Field(min_length=1, max_length=120)
    chart: ChartRequest
    annotation: str = Field(default="", max_length=2_000)
    title: str = ""
    scope: str = ""
    evidence: list[str] = Field(default_factory=list, max_length=20)
    warnings: list[str] = Field(default_factory=list, max_length=20)
    # Immutable server-validated result retained as report evidence; never client supplied.
    result: ChartResult | None = None


class CustomReportDocument(BaseModel):
    """Durable, run-scoped authored briefing; evidence is always server-derived."""
    run_id: str
    title: str = Field(default="Custom Report", min_length=1, max_length=120)
    executive_summary: str = Field(default="", max_length=8_000)
    sections: list[ReportSection] = Field(default_factory=list, max_length=20)
    pinned_artifacts: list[CustomReportArtifact] = Field(default_factory=list, max_length=24)
    manual_glossary_notes: list[ManualGlossaryNote] = Field(default_factory=list, max_length=30)
    glossary: list[dict[str, str]] = Field(default_factory=list)
    updated_at: str = ""


class CustomReportUpdate(BaseModel):
    title: str = Field(default="Custom Report", min_length=1, max_length=120)
    executive_summary: str = Field(default="", max_length=8_000)
    sections: list[ReportSection] = Field(default_factory=list, max_length=20)
    pinned_artifacts: list[CustomReportArtifact] = Field(default_factory=list, max_length=24)
    manual_glossary_notes: list[ManualGlossaryNote] = Field(default_factory=list, max_length=30)


class PinArtifactRequest(BaseModel):
    artifact_id: str = Field(min_length=1, max_length=120)
    chart: ChartRequest
    annotation: str = Field(default="", max_length=2_000)

class TextFilterRequest(BaseModel):
    text: str = Field(min_length=3, max_length=600)


class JobRequest(BaseModel):
    run_id: str
    kind: Literal["profile", "report"]


class ChatRequest(BaseModel):
    message: str = Field(min_length=2, max_length=1000)
    context: str = Field(default="", max_length=1000)
    language: Literal["en", "vi"] = "en"


class ClarificationOption(BaseModel):
    column: str
    label: str
    reason: str
    role: Literal["metric", "dimension"]


class SemanticSelectionRequest(BaseModel):
    """A user-confirmed schema choice, retained only with its report run."""
    column: str = Field(min_length=1, max_length=160)
    role: Literal["metric", "dimension"]
    language: str = Field(default="en", min_length=2, max_length=12)


class SemanticSelection(BaseModel):
    column: str
    role: Literal["metric", "dimension"]


class ChatResponse(BaseModel):
    answer: str
    insight: str
    scope: str
    title: str = ""
    chart: "ChartResult | None" = None
    table: list[dict[str, str | float | int]] = Field(default_factory=list)
    caveats: list[str] = Field(default_factory=list)
    clarification_options: list[ClarificationOption] = Field(default_factory=list)
    proposals: list[dict[str, object]] = Field(default_factory=list)
    mode: Literal["analysis", "clarification"]
    planner: Literal["llm", "deterministic"] = "deterministic"


class ChartResult(BaseModel):
    dimension: str
    metric: str
    aggregation: str
    chart_type: str
    title: str
    metric_display_name: str = ""
    value_format: dict[str, object] = Field(default_factory=dict)
    secondary_dimension: str | None = None
    filters: list[FilterSpec] = Field(default_factory=list)
    rows: list[dict[str, str | float | int]]
    warnings: list[str]
    sort_mode: Literal["chronological", "ranking"] = "ranking"
    result_count: int = 0
    insight_headline: str = ""
    evidence: list[str] = Field(default_factory=list)


class ExecutiveOverview(BaseModel):
    run_id: str
    summary: str
    charts: list[ChartResult]
    warnings: list[str] = Field(default_factory=list)
    guardrail: str


def safe_name(name: str) -> str:
    return Path(name).name


def is_number(value: str) -> bool:
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def is_date(value: str) -> bool:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            datetime.strptime(value.strip(), fmt)
            return True
        except ValueError:
            continue
    return False


def infer_kind(name: str, values: list[str]) -> str:
    observed = [value.strip() for value in values if value.strip()]
    if not observed:
        return "unknown"
    lower = name.lower()
    numeric_ratio = sum(is_number(value) for value in observed) / len(observed)
    date_ratio = sum(is_date(value) for value in observed) / len(observed)
    distinct = len(set(observed))
    if date_ratio >= .9 or (any(token in lower for token in ("date", "day", "month", "year", "period")) and date_ratio >= .5):
        return "time"
    if numeric_ratio >= .95:
        return "num"
    if ("id" in lower or "code" in lower) and distinct / len(observed) >= .9:
        return "id"
    return "cat"


def column_description(name: str, kind: str) -> str:
    hints = {"sale date": "Date on which the sales transaction was recorded.", "net sales": "Sales revenue after discounts and deductions.", "quantity": "Quantity recorded for the transaction.", "gross margin": "Difference between sales revenue and cost of goods sold."}
    return hints.get(name.lower().replace("_", " "), f"Inferred {kind} field from column name and observed values.")


def validate_headers(headers: list[str]) -> list[str]:
    clean = [str(header or "").strip() for header in headers]
    if len(clean) < 2 or any(not header for header in clean) or len(clean) != len(set(clean)):
        raise HTTPException(422, "Dataset must have at least two present, unique column headers.")
    return clean


SENSITIVE_COLUMN_TOKENS = {"email", "e-mail", "phone", "mobile", "address", "password", "token", "secret", "ssn", "passport", "national_id", "credit_card"}


def is_sensitive_column(name: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return any(token in normalized for token in SENSITIVE_COLUMN_TOKENS)


def csv_rows(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    headers = validate_headers(reader.fieldnames or [])
    rows: list[dict[str, str]] = []
    for row_index, row in enumerate(reader, start=1):
        if row_index > MAX_PROFILE_ROWS:
            raise HTTPException(422, f"Dataset exceeds the {MAX_PROFILE_ROWS:,}-row first-release limit.")
        rows.append({header: (row.get(header) or "").strip() for header in headers})
    return headers, rows


def xlsx_rows(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True, keep_vba=False)
    except Exception as error:
        raise HTTPException(422, "Workbook could not be read. Password-protected or malformed XLSX files are not supported.") from error
    sheet = book.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = validate_headers([str(value or "") for value in next(iterator)])
    except StopIteration as error:
        raise HTTPException(422, "Workbook has no header row.") from error
    rows: list[dict[str, str]] = []
    for row_index, values in enumerate(iterator, start=1):
        if row_index > MAX_PROFILE_ROWS:
            raise HTTPException(422, f"Dataset exceeds the {MAX_PROFILE_ROWS:,}-row first-release limit.")
        row = {headers[index]: "" if index >= len(values) or values[index] is None else str(values[index]).strip() for index in range(len(headers))}
        if any(row.values()):
            rows.append(row)
    return headers, rows


def persist(run_id: str, headers: list[str], rows: list[dict[str, str]], *, file_name: str, source_type: str = "file", source_label: str | None = None) -> None:
    RUN_STORE.save_dataset(run_id, headers, rows, file_name=file_name, source_type=source_type, source_label=source_label)


def load_run(run_id: str) -> tuple[list[str], list[dict[str, str]]]:
    headers, rows = RUN_STORE.load_dataset(run_id)
    return validate_headers(headers), rows


def build_profile(file_name: str, headers: list[str], rows: list[dict[str, str]], *, run_id: str, sample_limit: int | None = None) -> DatasetProfile:
    """Build a safe profile from all rows or a deterministic bounded prefix."""
    if not rows:
        raise HTTPException(422, "Dataset does not contain data rows.")
    if len(rows) > MAX_PROFILE_ROWS:
        raise HTTPException(422, f"Dataset exceeds the {MAX_PROFILE_ROWS:,}-row first-release limit.")
    profiled_rows = rows if sample_limit is None else rows[:sample_limit]
    profiles: list[ColumnProfile] = []
    warnings: list[str] = []
    for header in headers:
        values = [(row.get(header) or "").strip() for row in profiled_rows]
        null_count = sum(not value for value in values)
        kind = "id" if is_sensitive_column(header) else infer_kind(header, values)
        description = "Sensitive field; values are masked and cannot be used in analysis." if is_sensitive_column(header) else column_description(header, kind)
        profiles.append(ColumnProfile(name=header, kind=cast(Literal["time", "num", "cat", "id", "unknown"], kind), null_count=null_count, null_ratio=round(null_count / len(profiled_rows), 4), distinct_count=len({value for value in values if value}), description=description))
        if null_count / len(profiled_rows) >= .95:
            warnings.append(f"{header} is {null_count / len(profiled_rows):.1%} empty and may not be useful for charts.")
    units = next((item for item in profiles if item.name.lower() in {"unit_of_measure", "uom"}), None)
    quantity = next((item for item in profiles if item.name.lower() == "quantity"), None)
    if units and quantity and units.distinct_count > 1:
        warnings.append("Quantity has multiple units of measure; do not sum it until a compatible unit filter is applied.")
    sampled = len(profiled_rows) < len(rows)
    if sampled:
        warnings.append(f"Column metadata is sampled from the first {len(profiled_rows):,} rows; request full analysis for complete statistics.")
    return DatasetProfile(run_id=run_id, file_name=file_name, row_count=len(rows), column_count=len(headers), usable_column_count=sum(item.kind != "unknown" for item in profiles), columns=profiles, warnings=warnings, preview=safe_preview(rows), profile_status="sampled" if sampled else "complete", profiled_row_count=len(profiled_rows))


def profile_for_run(run_id: str, headers: list[str], rows: list[dict[str, str]]) -> DatasetProfile:
    """Compute the full profile only on demand, then retain it per API process."""
    cached = PROFILE_CACHE.get(run_id)
    if cached is not None:
        return cached
    metadata = RUN_STORE.metadata(run_id)
    result = build_profile(str(metadata["file_name"]), headers, rows, run_id=run_id)
    PROFILE_CACHE[run_id] = result
    return result


def profile(file_name: str, headers: list[str], rows: list[dict[str, str]], persist_run: bool = True, run_id: str | None = None, source_type: str = "file", source_label: str | None = None) -> DatasetProfile:
    """Create a complete profile for callers that already require rich metadata."""
    resolved_run_id = run_id or uuid4().hex
    if persist_run:
        persist(resolved_run_id, headers, rows, file_name=file_name, source_type=source_type, source_label=source_label)
    result = build_profile(file_name, headers, rows, run_id=resolved_run_id)
    PROFILE_CACHE[resolved_run_id] = result
    return result


def attach_profile(file_name: str, headers: list[str], rows: list[dict[str, str]], *, source_type: str = "file", source_label: str | None = None) -> DatasetProfile:
    """Persist a validated run first, then return only its bounded attach profile."""
    if not rows:
        raise HTTPException(422, "Dataset does not contain data rows.")
    if len(rows) > MAX_PROFILE_ROWS:
        raise HTTPException(422, f"Dataset exceeds the {MAX_PROFILE_ROWS:,}-row first-release limit.")
    run_id = uuid4().hex
    persist(run_id, headers, rows, file_name=file_name, source_type=source_type, source_label=source_label)
    return build_profile(file_name, headers, rows, run_id=run_id, sample_limit=ATTACH_PROFILE_SAMPLE_ROWS)


def safe_preview(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [{column: "[masked]" if is_sensitive_column(column) and value else value for column, value in row.items()} for row in rows[:20]]


def _finite_number(value: str) -> float | None:
    """Return a JSON-safe numeric value, without accepting NaN or infinity."""
    try:
        number = float(value.replace(",", ""))
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _numeric_summary(values: list[str]) -> dict[str, float | int]:
    numbers = sorted(number for value in values if (number := _finite_number(value)) is not None)
    count = len(numbers)
    if not count:
        return {"valid_count": 0, "invalid_count": len(values)}
    midpoint = count // 2
    median = numbers[midpoint] if count % 2 else (numbers[midpoint - 1] + numbers[midpoint]) / 2
    return {
        "valid_count": count,
        "invalid_count": len(values) - count,
        "min": numbers[0],
        "max": numbers[-1],
        "mean": round(sum(numbers) / count, 6),
        "median": round(median, 6),
    }


def _time_coverage(values: list[str]) -> dict[str, str | int]:
    parsed = sorted(value for value in (parse_date_value(item) for item in values) if value is not None)
    if not parsed:
        return {"valid_count": 0, "invalid_count": len(values)}
    return {
        "valid_count": len(parsed),
        "invalid_count": len(values) - len(parsed),
        "start": parsed[0].isoformat(),
        "end": parsed[-1].isoformat(),
    }


def _top_categories(values: list[str]) -> list[dict[str, str | int]]:
    counts = Counter(values)
    return [
        {"value": value, "count": count}
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:MAX_EDA_TOP_CATEGORIES]
    ]


def exploratory_data_analysis(run_id: str, headers: list[str], rows: list[dict[str, str]]) -> dict[str, object]:
    """Produce a bounded, deterministic run-scoped EDA summary.

    This deliberately avoids database query input and skips all value-level output
    for sensitive columns. Category values are capped and only returned for fields
    whose names pass the existing sensitive-column guard.
    """
    metadata = RUN_STORE.metadata(run_id)
    profile_data = profile_for_run(run_id, headers, rows)
    profiles = {item.name: item for item in profile_data.columns}
    visible_headers = [header for header in headers if not is_sensitive_column(header)][:MAX_EDA_COLUMNS]
    columns: list[dict[str, object]] = []
    for header in visible_headers:
        values = [(row.get(header) or "").strip() for row in rows]
        observed = [value for value in values if value]
        column = profiles[header]
        quality: dict[str, object] = {
            "non_null_count": len(observed),
            "null_count": column.null_count,
            "null_ratio": column.null_ratio,
            "distinct_count": column.distinct_count,
            "distinct_ratio": round(column.distinct_count / len(observed), 4) if observed else 0.0,
        }
        detail: dict[str, object] = {}
        if column.kind == "num":
            detail["numeric_summary"] = _numeric_summary(observed)
        elif column.kind == "time":
            detail["time_coverage"] = _time_coverage(observed)
        elif column.kind == "cat":
            detail["top_categories"] = _top_categories(observed)
        columns.append({"name": header, "kind": column.kind, "quality": quality, **detail})
    sensitive_column_count = sum(is_sensitive_column(header) for header in headers)
    return {
        "run_id": run_id,
        "coverage": {
            "row_count": len(rows),
            "column_count": len(headers),
            "analyzed_column_count": len(columns),
            "suppressed_sensitive_column_count": sensitive_column_count,
            "suppressed_column_count": len(headers) - len(columns) - sensitive_column_count,
        },
        "columns": columns,
        "provenance": {
            "dataset_sha256": hashlib.sha256(RUN_STORE.dataset_path(run_id).read_bytes()).hexdigest(),
            "source_type": metadata["source_type"],
            "source_label": metadata["source_label"],
            "analysis": "deterministic, bounded run-scoped summary",
            "top_category_limit": MAX_EDA_TOP_CATEGORIES,
        },
        "guardrails": [
            "Sensitive columns are excluded from value-level EDA.",
            f"Top categories are limited to {MAX_EDA_TOP_CATEGORIES} per non-sensitive categorical column.",
            f"At most {MAX_EDA_COLUMNS} non-sensitive columns are analyzed.",
        ],
    }


def quote_identifier(name: str, headers: list[str]) -> str:
    if name not in headers:
        raise HTTPException(422, f"Unknown column: {name}")
    if is_sensitive_column(name):
        raise HTTPException(422, "Sensitive columns cannot be used in filters, previews, or charts.")
    return '"' + name.replace('"', '""') + '"'

def _data_query_from_params(page: int, page_size: int, search: str, sort_by: str | None, sort_direction: str, filters: str) -> DataQuery:
    """Decode the one JSON query parameter without accepting expressions or SQL."""
    try:
        decoded_filters = json.loads(filters)
    except json.JSONDecodeError as error:
        raise HTTPException(422, "filters must be a JSON array of validated filter objects.") from error
    if not isinstance(decoded_filters, list):
        raise HTTPException(422, "filters must be a JSON array of validated filter objects.")
    try:
        return DataQuery(page=page, page_size=page_size, search=search, sort_by=sort_by, sort_direction=sort_direction, filters=decoded_filters)
    except ValidationError as error:
        raise HTTPException(422, "Invalid data query.") from error


def _data_where(query: DataQuery, headers: list[str], visible_headers: list[str]) -> tuple[str, list[str | float]]:
    """Build only parameterized predicates against non-sensitive, non-identifier fields."""
    clauses: list[str] = []
    parameters: list[str | float] = []
    operators = {"equals": "=", "not_equals": "<>", "greater_than": ">", "greater_or_equal": ">=", "less_than": "<", "less_or_equal": "<="}
    for item in query.filters:
        if item.column not in visible_headers:
            raise HTTPException(422, "Only non-sensitive, non-identifier columns can be used in data exploration.")
        field = quote_identifier(item.column, headers)
        operator = operators[item.operator]
        if item.operator in {"greater_than", "greater_or_equal", "less_than", "less_or_equal"}:
            if not is_number(item.value):
                raise HTTPException(422, f"Numeric comparison requires a numeric value for {item.column}.")
            clauses.append(f"TRY_CAST(REPLACE({field}, ',', '') AS DOUBLE) {operator} ?")
            parameters.append(float(item.value.replace(",", "")))
        else:
            clauses.append(f"{field} {operator} ?")
            parameters.append(item.value)
    term = query.search.strip()
    if term:
        # Search visible columns only, so a sensitive value can neither match nor leak.
        clauses.append("(" + " OR ".join(f"LOWER(COALESCE({quote_identifier(header, headers)}, '')) LIKE ?" for header in visible_headers) + ")")
        parameters.extend([f"%{term.lower()}%"] * len(visible_headers))
    return (" WHERE " + " AND ".join(clauses)) if clauses else "", parameters


def _run_data_query(run_id: str, query: DataQuery) -> tuple[list[str], list[dict[str, object]], int, list[dict[str, str]]]:
    headers, rows = load_run(run_id)
    profile_columns = {item.name: item for item in profile_for_run(run_id, headers, rows).columns}
    visible_headers = [header for header in headers if not is_sensitive_column(header) and profile_columns[header].kind != "id"]
    if not visible_headers:
        raise HTTPException(422, "This run has no non-sensitive, non-identifier columns available for exploration.")
    sort_column = query.sort_by or visible_headers[0]
    sort_identifier = quote_identifier(sort_column, headers)
    if sort_column not in visible_headers:
        raise HTTPException(422, "Sensitive columns cannot be used in filters, previews, or charts.")
    where_clause, parameters = _data_where(query, headers, visible_headers)
    select_columns = ", ".join(quote_identifier(header, headers) for header in visible_headers)
    # Add all visible columns as stable secondary keys so pagination has no ambiguous ties.
    tie_breakers = ", ".join(f"{quote_identifier(header, headers)} ASC" for header in visible_headers if header != sort_column)
    order_by = f"{sort_identifier} {query.sort_direction.upper()}" + (f", {tie_breakers}" if tie_breakers else "")
    connection = duckdb.connect(":memory:")
    try:
        connection.execute("CREATE TABLE dataset AS SELECT * FROM read_csv_auto(?, all_varchar=true)", [str(RUN_STORE.dataset_path(run_id))])
        total = int(connection.execute(f"SELECT COUNT(*) FROM dataset{where_clause}", parameters).fetchone()[0])
        result = connection.execute(f"SELECT {select_columns} FROM dataset{where_clause} ORDER BY {order_by} LIMIT ? OFFSET ?", [*parameters, query.page_size, (query.page - 1) * query.page_size])
        rows_out = [dict(zip(visible_headers, record, strict=True)) for record in result.fetchall()]
    finally:
        connection.close()
    columns = [{"name": header, "display_name": display_label(header), "kind": profile_columns[header].kind} for header in visible_headers]
    return visible_headers, rows_out, total, columns


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/readiness")
def readiness() -> dict[str, object]:
    """Readiness is intentionally dependency-light: sources are checked on use."""
    try:
        RUN_STORE.root.mkdir(parents=True, exist_ok=True)
        JOB_QUEUE.root.mkdir(parents=True, exist_ok=True)
        RUN_STORE.cleanup_expired()
    except OSError as error:
        raise HTTPException(503, "Artifact storage is not writable.") from error
    return {"status": "ready", "registered_source_count": len(registered_sources())}


@app.delete("/api/runs/{run_id}", status_code=204)
def delete_run(run_id: str) -> None:
    """Explicitly delete a run and every retained artifact; idempotency is not assumed."""
    RUN_STORE.metadata(run_id)
    import shutil
    shutil.rmtree(RUN_STORE._dir(run_id), ignore_errors=True)
    PROFILE_CACHE.pop(run_id, None)


@app.post("/api/maintenance/cleanup")
def cleanup_expired_runs(request: Request) -> dict[str, int]:
    """Internal scheduler hook; disabled unless an operator configures its key."""
    key = os.getenv("OPENDATA_MAINTENANCE_KEY", "")
    provided = request.headers.get("X-OpenData-Maintenance-Key", "")
    if not key or not secrets.compare_digest(provided, key):
        raise HTTPException(404, "Not found.")
    return {"removed_runs": RUN_STORE.cleanup_expired()}


@app.post("/api/jobs", status_code=202)
def enqueue_job(request: JobRequest) -> dict[str, object]:
    RUN_STORE.metadata(request.run_id)
    job_id = uuid4().hex
    return JOB_QUEUE.create(job_id, request.run_id, request.kind)


@app.get("/api/jobs/{job_id}")
def job_status(job_id: str) -> dict[str, object]:
    return JOB_QUEUE.get(job_id)


@app.delete("/api/jobs/{job_id}")
def cancel_job(job_id: str) -> dict[str, object]:
    return JOB_QUEUE.cancel(job_id)


@app.get("/api/sources")
def list_sources() -> dict[str, list[dict[str, object]]]:
    """Expose only operator-registered source metadata; never connection material."""
    return {"sources": [public_source(source) for source in registered_sources().values()]}


@app.post("/api/sources/{source_id}/runs", response_model=DatasetProfile, status_code=201)
def stage_registered_source(source_id: str) -> DatasetProfile:
    sources = registered_sources()
    source = sources.get(source_id)
    if source is None:
        raise HTTPException(404, "Registered source was not found.")
    headers, rows = read_registered_source(source)
    headers = validate_headers(headers)
    normalized = [{header: (row.get(header) or "").strip() for header in headers} for row in rows]
    if len(normalized) >= source.max_rows:
        raise HTTPException(422, f"Registered source reached its {source.max_rows:,}-row scan cap; narrow its operator configuration.")
    return profile(f"{source.display_name} ({source.locator})", headers, normalized, source_type=source.engine, source_label=f"{source.display_name} ({source.locator})")


def ingest_dataset(file_name: str, raw: bytes) -> DatasetProfile:
    """Validate and persist a file, returning a bounded profile for immediate attach."""
    name = safe_name(file_name or "upload")
    if not re.fullmatch(r"[\w .()\-]+\.(csv|xlsx)", name, flags=re.IGNORECASE):
        raise HTTPException(415, "Upload a CSV or XLSX file with a safe file name.")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File exceeds the {MAX_UPLOAD_BYTES // 1024 // 1024} MB upload limit.")
    headers, rows = xlsx_rows(raw) if name.lower().endswith(".xlsx") else csv_rows(raw)
    return attach_profile(name, headers, rows)


@app.post("/api/runs/upload", response_model=DatasetProfile, status_code=201)
async def upload_dataset(file: UploadFile = File(...)) -> DatasetProfile:
    return ingest_dataset(file.filename or "upload", await file.read(MAX_UPLOAD_BYTES + 1))


@app.get("/api/runs/{run_id}/profile/status", response_model=DatasetProfile)
def profile_status(run_id: str) -> DatasetProfile:
    """Lazily materialize the complete profile; later requests use the process cache."""
    cached = PROFILE_CACHE.get(run_id)
    if cached is not None:
        return cached
    headers, rows = load_run(run_id)
    return profile_for_run(run_id, headers, rows)


@app.get("/api/runs/{run_id}/plan")
def suggested_plan(run_id: str, limit: int = 8) -> dict[str, object]:
    headers, rows = load_run(run_id)
    profile_data = profile_for_run(run_id, headers, rows)
    return {"charts": propose_charts(profile_data.columns, max(1, min(limit, 12))), "note": "Candidates are deterministic suggestions; review and approve before report generation."}


def starter_views_payload(run_id: str, language: Literal["en", "vi"] = "en") -> dict[str, object]:
    """Return safe, selectable views from profile metadata only; never execute a chart."""
    headers, rows = load_run(run_id)
    profile_data = profile_for_run(run_id, headers, rows)
    proposals = analyst_proposals(profile_data.columns, language=language)
    for proposal in proposals:
        chart_request = ChartRequest.model_validate(cast(dict[str, object], proposal["request"]))
        # The card payload is the exact schema-validated request that can execute.
        proposal["request"] = chart_request.model_dump()
        proposal["question"] = (f"Tổng {chart_request.metric} theo {chart_request.dimension}" if language == "vi" else f"Show sum of {chart_request.metric} by {chart_request.dimension}")
        proposal["prompt"] = proposal["question"]

    return {
        "summary": f"Run này có {profile_data.row_count:,} dòng, {profile_data.usable_column_count} cột có thể dùng, {sum(item.kind == 'num' for item in profile_data.columns)} chỉ tiêu số và {sum(item.kind in {'cat', 'time'} for item in profile_data.columns)} trường phân tích." if language == "vi" else f"This run has {profile_data.row_count:,} rows, {profile_data.usable_column_count} usable columns, {sum(item.kind == 'num' for item in profile_data.columns)} metrics and {sum(item.kind in {'cat', 'time'} for item in profile_data.columns)} dimensions/time fields.",
        "proposals": proposals,
        "guardrail": "Gợi ý chỉ dùng vai trò cột suy luận và số liệu profile. Chart chỉ chạy sau khi bạn phê duyệt và luôn dùng aggregate đã xác thực trên server." if language == "vi" else "Suggestions use inferred column roles and profile counts only. Charts run only after your approval and always use validated server-side aggregates.",
    }


@app.get("/api/runs/{run_id}/analyst-proposals")
def analyst_plan(run_id: str, language: Literal["en", "vi"] = "en") -> dict[str, object]:
    return starter_views_payload(run_id, language)


@app.get("/api/runs/{run_id}/starter-views")
def starter_views(run_id: str, language: Literal["en", "vi"] = "en") -> dict[str, object]:
    return starter_views_payload(run_id, language)


@app.get("/api/runs/{run_id}/eda")
def eda_for_run(run_id: str) -> dict[str, object]:
    """Return deterministic, bounded EDA for one retained report run."""
    headers, rows = load_run(run_id)
    return exploratory_data_analysis(run_id, headers, rows)


@app.post("/api/runs/{run_id}/parse-filter")
def parse_text_filter(run_id: str, request: TextFilterRequest) -> dict[str, object]:
    headers, _ = load_run(run_id)
    try:
        parsed = parse_filter(request.text, headers)
    except ValueError as error:
        raise HTTPException(422, str(error)) from error
    return {"filter": {"column": parsed.column, "operator": parsed.operator, "value": parsed.value}, "confirmation": f"Apply {parsed.column} {parsed.operator.replace('_', ' ')} {parsed.value}?"}


@app.get("/api/runs/{run_id}/data")
def explore_data(run_id: str, page: int = 1, page_size: int = 25, search: str = "", sort_by: str | None = None, sort_direction: Literal["asc", "desc"] = "asc", filters: str = "[]") -> dict[str, object]:
    query = _data_query_from_params(page, page_size, search, sort_by, sort_direction, filters)
    _, result_rows, total, columns = _run_data_query(run_id, query)
    page_count = max(1, math.ceil(total / query.page_size))
    return {"run_id": run_id, "columns": columns, "rows": result_rows, "total": total, "filters": [item.model_dump() for item in query.filters], "pagination": {"page": query.page, "page_size": query.page_size, "page_count": page_count, "has_next": query.page < page_count, "has_previous": query.page > 1}}


@app.get("/api/runs/{run_id}/data/export")
def export_data(run_id: str, page: int = 1, page_size: int = 25, search: str = "", sort_by: str | None = None, sort_direction: Literal["asc", "desc"] = "asc", filters: str = "[]") -> StreamingResponse:
    """Export the active filtered scope only; reject, rather than truncate, oversized exports."""
    query = _data_query_from_params(page, page_size, search, sort_by, sort_direction, filters)
    export_query = query.model_copy(update={"page": 1, "page_size": MAX_DATA_EXPORT_ROWS})
    headers, result_rows, total, _ = _run_data_query(run_id, export_query)
    if total > MAX_DATA_EXPORT_ROWS:
        raise HTTPException(422, f"Filtered export has {total:,} rows; refine filters to at most {MAX_DATA_EXPORT_ROWS:,} rows before exporting.")
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    # Prevent spreadsheet applications from evaluating formula-leading cells on open.
    safe_rows = [{key: (f"'{value}" if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"} else value) for key, value in row.items()} for row in result_rows]
    writer.writerows(safe_rows)
    filename = f"opendata-{run_id[:8]}-filtered.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"', "X-OpenData-Export-Row-Count": str(total)})


@app.get("/api/runs/{run_id}/values/{column}")
def values(run_id: str, column: str) -> dict[str, str | list[str]]:
    headers, rows = load_run(run_id)
    quote_identifier(column, headers)
    result = sorted({(row.get(column) or "").strip() for row in rows if (row.get(column) or "").strip()})[:100]
    return {"column": column, "values": result}


@app.post("/api/runs/{run_id}/chart", response_model=ChartResult)
def build_chart(run_id: str, request: ChartRequest, language: Literal["en", "vi"] = "vi") -> ChartResult:
    headers, rows = load_run(run_id)
    dimension = quote_identifier(request.dimension, headers)
    metric = quote_identifier(request.metric, headers)
    secondary = quote_identifier(request.secondary_dimension, headers) if request.secondary_dimension else None
    x_metric = quote_identifier(request.x_metric, headers) if request.x_metric else None
    profile_data = profile_for_run(run_id, headers, rows)
    profiles = {item.name: item for item in profile_data.columns}
    dimension_profile = profiles[request.dimension]
    metric_profile = profiles[request.metric]
    if metric_profile.kind != "num" and request.aggregation != "count":
        raise HTTPException(422, "The selected metric must be numeric.")
    if request.chart_type in {"stacked_bar", "heatmap"} and not secondary:
        raise HTTPException(422, f"{request.chart_type} requires a secondary_dimension.")
    if request.chart_type in {"pie", "donut"} and (secondary or dimension_profile.kind != "cat"):
        raise HTTPException(422, f"{request.chart_type} requires one categorical dimension and no secondary_dimension.")
    if request.chart_type == "scatter":
        if secondary or not request.x_metric or request.x_metric == request.metric or profiles[request.x_metric].kind != "num":
            raise HTTPException(422, "scatter requires distinct numeric metric and x_metric fields and no secondary_dimension.")
    if request.aggregation == "count": expression = "COUNT(*)"
    else: expression = f"{request.aggregation.upper()}(TRY_CAST(REPLACE({metric}, ',', '') AS DOUBLE))"
    x_expression = f"{request.aggregation.upper()}(TRY_CAST(REPLACE({x_metric}, ',', '') AS DOUBLE))" if x_metric else None
    filter_clauses = [f"{dimension} IS NOT NULL", f"TRIM({dimension}) <> ''"]
    if secondary:
        filter_clauses.extend([f"{secondary} IS NOT NULL", f"TRIM({secondary}) <> ''"])
    parameters: list[str | int | float] = []
    for item in request.filters:
        field = quote_identifier(item.column, headers)
        operators = {"equals": "=", "not_equals": "<>", "greater_than": ">", "greater_or_equal": ">=", "less_than": "<", "less_or_equal": "<="}
        operator = operators[item.operator]
        if item.operator in {"greater_than", "greater_or_equal", "less_than", "less_or_equal"}:
            if not is_number(item.value):
                raise HTTPException(422, f"Numeric comparison requires a numeric value for {item.column}.")
            filter_clauses.append(f"TRY_CAST(REPLACE({field}, ',', '') AS DOUBLE) {operator} ?")
            parameters.append(float(item.value.replace(",", "")))
        else:
            filter_clauses.append(f"{field} {operator} ?")
            parameters.append(item.value)
    chronological = request.chart_type in {"line", "area"} and dimension_profile.kind == "time" and not secondary
    connection = duckdb.connect(":memory:")
    try:
        connection.execute("CREATE TABLE dataset AS SELECT * FROM read_csv_auto(?, all_varchar=true)", [str(RUN_STORE.dataset_path(run_id))])
        where_clause = " AND ".join(filter_clauses)
        if request.chart_type == "scatter":
            query = f"SELECT {dimension} AS label, {x_expression} AS x_value, {expression} AS value FROM dataset WHERE {where_clause} GROUP BY 1 ORDER BY value DESC NULLS LAST LIMIT ?"
        elif secondary and request.limit_per_secondary:
            query = f"""WITH aggregates AS (
                SELECT {dimension} AS label, {secondary} AS secondary_label, {expression} AS value
                FROM dataset WHERE {where_clause} GROUP BY 1, 2
            ), ranked AS (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY secondary_label ORDER BY value DESC NULLS LAST, label ASC) AS group_rank
                FROM aggregates
            )
            SELECT label, secondary_label, value FROM ranked
            WHERE group_rank <= ? ORDER BY secondary_label ASC, value DESC NULLS LAST, label ASC"""
        elif secondary:
            query = f"SELECT {dimension} AS label, {secondary} AS secondary_label, {expression} AS value FROM dataset WHERE {where_clause} GROUP BY 1, 2 ORDER BY value DESC NULLS LAST LIMIT ?"
        elif chronological:
            query = f"SELECT {dimension} AS label, {expression} AS value FROM dataset WHERE {where_clause} GROUP BY 1 ORDER BY TRY_CAST({dimension} AS TIMESTAMP) ASC NULLS LAST LIMIT ?"
        else:
            query = f"SELECT {dimension} AS label, {expression} AS value FROM dataset WHERE {where_clause} GROUP BY 1 ORDER BY value DESC NULLS LAST LIMIT ?"
        records = connection.execute(query, [*parameters, request.limit]).fetchall()
    finally:
        connection.close()
    warnings: list[str] = []
    if not records: warnings.append("Không có giá trị phù hợp với phạm vi hiện tại." if language == "vi" else "No values match the current scope.")
    if request.metric.lower() == "quantity" and request.aggregation == "sum": warnings.append("Cần lọc theo đơn vị tương thích trước khi diễn giải tổng quantity." if language == "vi" else "Filter to compatible units before interpreting a total quantity.")
    if request.chart_type in {"line", "area"} and not chronological:
        warnings.append("Trục đang không phải trường thời gian đã xác nhận; kết quả được xếp hạng theo giá trị thay vì diễn giải như xu hướng." if language == "vi" else "The axis is not a confirmed time field; results are ranked by value rather than interpreted as a trend.")
    if request.chart_type == "scatter":
        chart_rows = [{"label": str(label), "display_label": format_display_date(str(label)) if dimension_profile.kind == "time" else str(label), "x_value": 0 if x_value is None else float(x_value), "value": 0 if value is None else float(value), "formatted_value": format_number(0 if value is None else float(value))} for label, x_value, value in records]
    elif secondary:
        chart_rows = [{"label": str(label), "display_label": format_display_date(str(label)) if dimension_profile.kind == "time" else str(label), "secondary_label": str(second), "value": 0 if value is None else float(value), "formatted_value": format_number(0 if value is None else float(value))} for label, second, value in records]
    else:
        chart_rows = [{"label": str(label), "display_label": format_display_date(str(label)) if dimension_profile.kind == "time" else str(label), "value": 0 if value is None else float(value), "formatted_value": format_number(0 if value is None else float(value))} for label, value in records]
    if request.chart_type == "pareto":
        total = sum(float(row["value"]) for row in chart_rows)
        running = 0.0
        for row in chart_rows:
            running += float(row["value"])
            row["cumulative_pct"] = 0 if total == 0 else round(running / total * 100, 2)
    insight_headline, evidence = chart_insight(chart_rows, chronological, language)
    per_secondary = bool(secondary and request.limit_per_secondary)
    label = ("Xu hướng" if chronological else f"Top {request.limit if per_secondary else len(chart_rows)}") if language == "vi" else ("Trend" if chronological else f"Top {request.limit if per_secondary else len(chart_rows)}")
    if per_secondary:
        title = f"{label} {display_label(request.metric)} theo {display_label(request.dimension)} trong mỗi {display_label(request.secondary_dimension or '')}" if language == "vi" else f"{label} {display_label(request.metric)} by {display_label(request.dimension)} per {display_label(request.secondary_dimension or '')}"
    else:
        title = (f"{label} {display_label(request.metric)} theo {display_label(request.dimension)}" if language == "vi" else f"{label} {display_label(request.metric)} by {display_label(request.dimension)}") + (f" × {display_label(request.secondary_dimension or '')}" if secondary else "")
    return ChartResult(dimension=request.dimension, metric=request.metric, aggregation=request.aggregation, chart_type=request.chart_type, title=title, metric_display_name=display_label(request.metric), value_format=value_format_descriptor(), secondary_dimension=request.secondary_dimension, filters=request.filters, rows=chart_rows, warnings=warnings, sort_mode="chronological" if chronological else "ranking", result_count=len(chart_rows), insight_headline=insight_headline, evidence=evidence)


@app.get("/api/runs/{run_id}/executive-overview", response_model=ExecutiveOverview)
def executive_overview(run_id: str, language: Literal["en", "vi"] = "en") -> ExecutiveOverview:
    """Generate a bounded, read-only overview from validated server aggregates."""
    headers, rows = load_run(run_id)
    profile_data = profile_for_run(run_id, headers, rows)
    charts: list[ChartResult] = []
    warnings: list[str] = []
    for proposal in analyst_proposals(profile_data.columns, max_charts=5, language=language):
        try:
            request = ChartRequest.model_validate(cast(dict[str, object], proposal["request"]))
            chart = build_chart(run_id, request, language)
            if chart.rows:
                charts.append(chart)
            else:
                warnings.append(f"{chart.title}: " + (chart.warnings[0] if chart.warnings else "no matching values"))
        except (HTTPException, ValueError, TypeError) as error:
            warnings.append(str(getattr(error, "detail", error))[:220])
    summary = (f"Bộ tổng quan gồm {len(charts)} biểu đồ aggregate đã xác thực từ {profile_data.row_count:,} dòng." if language == "vi" else f"This executive overview contains {len(charts)} validated aggregate charts from {profile_data.row_count:,} rows.")
    guardrail = ("Các biểu đồ chỉ dùng aggregate run-scoped đã xác thực trên server; không dùng raw rows hoặc trường nhạy cảm." if language == "vi" else "Charts use only validated, run-scoped server aggregates; no raw rows or sensitive fields are used.")
    return ExecutiveOverview(run_id=run_id, summary=summary, charts=charts, warnings=warnings, guardrail=guardrail)


def _custom_report_glossary(run_id: str, artifacts: list[CustomReportArtifact]) -> list[dict[str, str]]:
    headers, rows = load_run(run_id)
    profile = {item.name: item for item in profile_for_run(run_id, headers, rows).columns}
    used = {name for artifact in artifacts for name in (artifact.chart.dimension, artifact.chart.metric, artifact.chart.secondary_dimension) if name}
    return [{"name": name, "label": display_label(name), "description": profile[name].description, "kind": profile[name].kind} for name in sorted(used) if name in profile and not is_sensitive_column(name)]


def _report_artifact(run_id: str, artifact: CustomReportArtifact) -> CustomReportArtifact:
    """Rebuild the immutable report snapshot; client artifact metadata is never trusted."""
    # CustomReportUpdate accepts the complete document for editor convenience.  Only
    # the chart specification and author annotation are client-authored, however:
    # title, scope, evidence, warnings, and result must always be derived from this
    # run's data at write time.
    chart = build_chart(run_id, artifact.chart)
    scope = f"{chart.aggregation} of {display_label(chart.metric)} by {display_label(chart.dimension)}"
    if chart.secondary_dimension:
        scope += f"; grouped by {display_label(chart.secondary_dimension)}"
    if chart.filters:
        scope += "; filtered to " + "; ".join(f"{display_label(item.column)} {item.operator.replace('_', ' ')} {item.value}" for item in chart.filters)
    scope += f"; top {chart.result_count or len(chart.rows)} {chart.sort_mode or 'results'}."
    return CustomReportArtifact(artifact_id=artifact.artifact_id, chart=artifact.chart, annotation=artifact.annotation, title=chart.title, scope=scope, evidence=chart.evidence, warnings=chart.warnings, result=chart)


def _custom_report_document(run_id: str, update: CustomReportUpdate | None = None) -> CustomReportDocument:
    path = "custom-report.json"
    if update is None:
        try:
            return CustomReportDocument.model_validate(RUN_STORE.artifact_json(run_id, path))
        except HTTPException as error:
            if error.status_code != 404: raise
            return CustomReportDocument(run_id=run_id, glossary=[])
    artifacts = list({item.artifact_id: _report_artifact(run_id, item) for item in update.pinned_artifacts}.values())
    document = CustomReportDocument(run_id=run_id, title=update.title, executive_summary=update.executive_summary, sections=update.sections, pinned_artifacts=artifacts, manual_glossary_notes=update.manual_glossary_notes, glossary=_custom_report_glossary(run_id, artifacts), updated_at=datetime.now(timezone.utc).isoformat())
    RUN_STORE.save_artifact_json(run_id, path, document.model_dump())
    return document


@app.get("/api/runs/{run_id}/custom-report", response_model=CustomReportDocument)
def get_custom_report(run_id: str) -> CustomReportDocument:
    RUN_STORE.metadata(run_id)
    return _custom_report_document(run_id)


@app.put("/api/runs/{run_id}/custom-report", response_model=CustomReportDocument)
def update_custom_report(run_id: str, request: CustomReportUpdate) -> CustomReportDocument:
    return _custom_report_document(run_id, request)


@app.post("/api/runs/{run_id}/custom-report/artifacts", response_model=CustomReportDocument)
def pin_custom_report_artifact(run_id: str, request: PinArtifactRequest) -> CustomReportDocument:
    current = _custom_report_document(run_id)
    artifacts = [item for item in current.pinned_artifacts if item.artifact_id != request.artifact_id] + [CustomReportArtifact(artifact_id=request.artifact_id, chart=request.chart, annotation=request.annotation)]
    return _custom_report_document(run_id, CustomReportUpdate(title=current.title, executive_summary=current.executive_summary, sections=current.sections, pinned_artifacts=artifacts, manual_glossary_notes=current.manual_glossary_notes))


@app.delete("/api/runs/{run_id}/custom-report/artifacts/{artifact_id}", response_model=CustomReportDocument)
def unpin_custom_report_artifact(run_id: str, artifact_id: str) -> CustomReportDocument:
    current = _custom_report_document(run_id)
    artifacts = [item for item in current.pinned_artifacts if item.artifact_id != artifact_id]
    return _custom_report_document(run_id, CustomReportUpdate(title=current.title, executive_summary=current.executive_summary, sections=current.sections, pinned_artifacts=artifacts, manual_glossary_notes=current.manual_glossary_notes))


@app.get("/api/runs/{run_id}/manifest")
def get_manifest(run_id: str) -> dict[str, object]:
    return RUN_STORE.artifact_json(run_id, "report.manifest.json")


def _report_chart_svg(chart: ChartResult) -> str:
    """Render a deterministic, self-contained SVG from a validated chart snapshot."""
    esc = lambda value: html.escape(str(value), quote=True)
    rows = chart.rows[:30]
    width, height = 760, 360
    left, top, right, bottom = 66, 38, 24, 66
    plot_width, plot_height = width - left - right, height - top - bottom
    title = f"{chart.title} ({chart.chart_type} chart)"
    description = f"{chart.aggregation} of {chart.metric_display_name or display_label(chart.metric)} by {display_label(chart.dimension)}. " + "; ".join(
        f"{row.get('display_label') or row.get('label', '')}: {row.get('formatted_value') or row.get('value', 0)}"
        for row in rows
    )
    if not rows:
        return (f"<svg class='report-chart' viewBox='0 0 {width} {height}' role='img' aria-labelledby='chart-title chart-desc' "
                f"xmlns='http://www.w3.org/2000/svg'><title id='chart-title'>{esc(title)}</title><desc id='chart-desc'>{esc(description)}</desc>"
                f"<text x='{width / 2}' y='{height / 2}' text-anchor='middle'>No validated chart values are available.</text></svg>")

    def numeric(row: dict[str, str | float | int], key: str = "value") -> float:
        try:
            value = float(row.get(key, 0) or 0)
            return value if math.isfinite(value) else 0.0
        except (TypeError, ValueError):
            return 0.0

    def label(row: dict[str, str | float | int]) -> str:
        return str(row.get("display_label") or row.get("label") or "")

    svg = [f"<svg class='report-chart' viewBox='0 0 {width} {height}' role='img' aria-labelledby='chart-title chart-desc' xmlns='http://www.w3.org/2000/svg'>",
           f"<title id='chart-title'>{esc(title)}</title><desc id='chart-desc'>{esc(description)}</desc>",
           "<rect width='100%' height='100%' rx='10' fill='#f8fafc'/>"]
    values = [numeric(row) for row in rows]
    colors = ("#2563eb", "#0d9488", "#7c3aed", "#ea580c", "#db2777", "#0891b2", "#65a30d", "#ca8a04")

    if chart.chart_type in {"pie", "donut"}:
        positive = [max(0.0, value) for value in values]
        total = sum(positive)
        center_x, center_y, radius = width / 2, height / 2, min(plot_height / 2 - 12, 118)
        if total <= 0:
            svg.append(f"<text x='{center_x}' y='{center_y}' text-anchor='middle'>No positive values to chart.</text>")
        else:
            angle = -math.pi / 2
            for index, (row, value) in enumerate(zip(rows, positive)):
                sweep = value / total * math.tau
                end = angle + sweep
                x1, y1 = center_x + radius * math.cos(angle), center_y + radius * math.sin(angle)
                x2, y2 = center_x + radius * math.cos(end), center_y + radius * math.sin(end)
                large = 1 if sweep > math.pi else 0
                path = f"M {center_x:.1f} {center_y:.1f} L {x1:.1f} {y1:.1f} A {radius:.1f} {radius:.1f} 0 {large} 1 {x2:.1f} {y2:.1f} Z"
                svg.append(f"<path d='{path}' fill='{colors[index % len(colors)]}'><title>{esc(label(row))}: {esc(row.get('formatted_value', value))}</title></path>")
                angle = end
            if chart.chart_type == "donut":
                svg.append(f"<circle cx='{center_x}' cy='{center_y}' r='{radius * .52:.1f}' fill='#f8fafc'/><text x='{center_x}' y='{center_y + 5}' text-anchor='middle' font-weight='700'>{esc(format_number(total))}</text>")
            for index, row in enumerate(rows[:8]):
                y = top + index * 24
                svg.append(f"<rect x='{width - 205}' y='{y - 10}' width='12' height='12' fill='{colors[index % len(colors)]}'/><text x='{width - 188}' y='{y}' font-size='12'>{esc(label(row)[:24])}: {esc(row.get('formatted_value', values[index]))}</text>")
    elif chart.chart_type == "scatter":
        x_values = [numeric(row, "x_value") for row in rows]
        x_min, x_max = min(x_values), max(x_values)
        y_min, y_max = min(values), max(values)
        x_span, y_span = max(1.0, x_max - x_min), max(1.0, y_max - y_min)
        svg.append(f"<path d='M {left} {top} V {height - bottom} H {width - right}' stroke='#64748b' fill='none'/>")
        for index, (row, x_value, value) in enumerate(zip(rows, x_values, values)):
            x = left + (x_value - x_min) / x_span * plot_width
            y = top + (y_max - value) / y_span * plot_height
            svg.append(f"<circle cx='{x:.1f}' cy='{y:.1f}' r='5' fill='{colors[index % len(colors)]}'><title>{esc(label(row))}: {esc(row.get('formatted_value', value))}</title></circle>")
        svg.append(f"<text x='{left}' y='{height - 18}' font-size='12'>{esc(format_number(x_min))}</text><text x='{width - right}' y='{height - 18}' text-anchor='end' font-size='12'>{esc(format_number(x_max))}</text>")
    else:
        minimum, maximum = min(0.0, min(values)), max(0.0, max(values))
        span = max(1.0, maximum - minimum)
        baseline = top + (maximum / span) * plot_height
        svg.append(f"<path d='M {left} {baseline:.1f} H {width - right}' stroke='#64748b' fill='none'/>")
        points: list[tuple[float, float]] = []
        step = plot_width / max(1, len(rows))
        for index, (row, value) in enumerate(zip(rows, values)):
            x = left + step * (index + .5)
            y = top + (maximum - value) / span * plot_height
            points.append((x, y))
            if chart.chart_type == "bar":
                bar_width = max(5, step * .68)
                bar_y, bar_height = min(y, baseline), abs(baseline - y)
                svg.append(f"<rect x='{x - bar_width / 2:.1f}' y='{bar_y:.1f}' width='{bar_width:.1f}' height='{max(1, bar_height):.1f}' rx='2' fill='{colors[index % len(colors)]}'><title>{esc(label(row))}: {esc(row.get('formatted_value', value))}</title></rect>")
            svg.append(f"<text x='{x:.1f}' y='{height - bottom + 18}' text-anchor='middle' font-size='11'>{esc(label(row)[:14])}</text>")
        if chart.chart_type in {"line", "area"}:
            point_string = " ".join(f"{x:.1f},{y:.1f}" for x, y in points)
            if chart.chart_type == "area":
                svg.append(f"<path d='M {points[0][0]:.1f},{baseline:.1f} L {point_string} L {points[-1][0]:.1f},{baseline:.1f} Z' fill='#93c5fd' opacity='.7'/>")
            svg.append(f"<polyline points='{point_string}' fill='none' stroke='#2563eb' stroke-width='3'/>")
            for (row, value), (x, y) in zip(zip(rows, values), points):
                svg.append(f"<circle cx='{x:.1f}' cy='{y:.1f}' r='4' fill='#1d4ed8'><title>{esc(label(row))}: {esc(row.get('formatted_value', value))}</title></circle>")
        svg.append(f"<text x='{left - 8}' y='{top + 4}' text-anchor='end' font-size='12'>{esc(format_number(maximum))}</text><text x='{left - 8}' y='{height - bottom + 4}' text-anchor='end' font-size='12'>{esc(format_number(minimum))}</text>")
    svg.append("</svg>")
    return "".join(svg)


@app.get("/api/runs/{run_id}/report", response_class=HTMLResponse)
@app.post("/api/runs/{run_id}/report", response_class=HTMLResponse)
def build_report(run_id: str, request: ReportRequest | None = None) -> HTMLResponse:
    """Render portable HTML from the persisted authored document, never client charts."""
    document = _custom_report_document(run_id)
    if request is not None and not document.pinned_artifacts:
        document = _custom_report_document(run_id, CustomReportUpdate(title=request.title, pinned_artifacts=[CustomReportArtifact(artifact_id=f"legacy-{index}", chart=chart) for index, chart in enumerate(request.charts)]))
    # Export exactly the persisted validated evidence; old artifacts are hydrated on read/save.
    charts = [item.result or build_chart(run_id, item.chart) for item in document.pinned_artifacts]
    evidence = [fact for chart in charts for fact in evidence_for_chart(chart)]
    metadata = RUN_STORE.metadata(run_id)
    manifest = {"run_id": run_id, "generated_at": datetime.now(timezone.utc).isoformat(), "dataset_sha256": hashlib.sha256(RUN_STORE.dataset_path(run_id).read_bytes()).hexdigest(), "source_type": metadata["source_type"], "source_label": metadata["source_label"], "chart_specs": [item.chart.model_dump() for item in document.pinned_artifacts], "chart_count": len(charts), "evidence": evidence, "document_updated_at": document.updated_at}
    RUN_STORE.save_artifact_json(run_id, "report.manifest.json", manifest)
    esc = lambda value: html.escape(str(value))
    artifact_parts = []
    for saved, chart in zip(document.pinned_artifacts, charts):
        rows = "".join("<tr><td>{}</td>{}<td>{}</td></tr>".format(esc(row.get("display_label") or row["label"]), "<td>{}</td>".format(esc(row.get("secondary_label") or "")) if chart.secondary_dimension else "", esc(row.get("formatted_value", row["value"]))) for row in chart.rows)
        artifact_parts.append("<section class='card'><h2>{}</h2><p class='scope'>{}</p>{}<div class='chart-visual'>{}</div><h3>Validated evidence</h3><ul>{}</ul>{}<table><caption>Accessible data table for {}</caption><thead><tr><th>{}</th>{}<th>{} {}</th></tr></thead><tbody>{}</tbody></table></section>".format(esc(saved.title or chart.title), esc(saved.scope), "<p><strong>Author note:</strong> {}</p>".format(esc(saved.annotation)) if saved.annotation else "", _report_chart_svg(chart), "".join("<li>{}</li>".format(esc(item)) for item in saved.evidence) or "<li>No summary evidence was available.</li>", esc(saved.title or chart.title), "".join("<p class='warning'>{}</p>".format(esc(item)) for item in saved.warnings), esc(display_label(chart.dimension)), "<th>{}</th>".format(esc(display_label(chart.secondary_dimension))) if chart.secondary_dimension else "", esc(chart.aggregation), esc(display_label(chart.metric)), rows))
    sections = "".join("<section class='card'><h2>{}</h2><p>{}</p>{}</section>".format(esc(section.heading), esc(section.commentary), "<h3>Recommended actions</h3><ul>{}</ul>".format("".join("<li>{}</li>".format(esc(action)) for action in section.recommended_actions)) if section.recommended_actions else "") for section in document.sections)
    glossary = "".join("<li><strong>{}</strong> ({}) — {}</li>".format(esc(item["label"]), esc(item["kind"]), esc(item["description"])) for item in document.glossary) or "<li>No validated glossary entries yet.</li>"
    notes = "".join("<li class='manual'><strong>Manual note:</strong> {}</li>".format(esc(note.text)) for note in document.manual_glossary_notes) or "<li class='manual'>No manual glossary notes.</li>"
    artifact = """<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{}</title><style>body{{font:15px system-ui;margin:0;background:#f8fafc;color:#172554}}main{{max-width:1100px;margin:auto;padding:36px}}.meta,.card{{background:#fff;border:1px solid #e2e8f0;border-radius:16px;padding:20px;margin:16px 0}}table{{width:100%;border-collapse:collapse}}th,td{{padding:9px;border-bottom:1px solid #e2e8f0;text-align:left;vertical-align:top}}.warning{{color:#92400e;background:#fffbeb;padding:10px;border-radius:8px}}.scope{{font-size:13px;color:#475569}}.manual{{color:#5b21b6}}.chart-visual{{overflow-x:auto;margin:16px 0}}.report-chart{{display:block;min-width:620px;width:100%;height:auto}}caption{{text-align:left;font-weight:600;padding:0 0 8px}}@media print{{body{{background:#fff}}main{{max-width:none;padding:0}}.card,.meta{{break-inside:avoid}}}}</style></head><body><main><h1>{}</h1><p>Authored briefing from validated report run <code>{}</code>. Use your browser’s Print command to save as PDF.</p><section class='meta'><h2>Executive summary</h2><p>{}</p></section>{}<section class='meta'><h2>Validated artifacts and evidence</h2>{}</section><section class='meta'><h2>Glossary</h2><ul>{}</ul><h3>Author notes (not validated evidence)</h3><ul>{}</ul></section><section class='meta'><h2>Provenance</h2><p>Dataset checksum: <code>{}</code>. Source: {} / {}. Generated: {}. Artifact specifications and evidence are retained in the run manifest.</p></section></main></body></html>""".format(esc(document.title), esc(document.title), esc(run_id[:8]), esc(document.executive_summary) or "No executive summary supplied.", sections, "".join(artifact_parts) or "<section class='card'><p>No validated artifacts have been pinned.</p></section>", glossary, notes, esc(manifest["dataset_sha256"]), esc(metadata["source_type"]), esc(metadata["source_label"]), esc(manifest["generated_at"]))
    compatibility_payload = json.dumps([chart.model_dump() for chart in charts]).replace("</", "<\\/")
    artifact += f"<!-- validated-artifact-json: {compatibility_payload} -->"
    return HTMLResponse(artifact, headers={"Content-Disposition": 'attachment; filename="opendata-authored-report.html"'})

def _llm_chart_request(columns: list[ColumnProfile], request: ChatRequest) -> tuple[ChartRequest | None, str | None]:
    """Ask the configured LLM for JSON intent only; never disclose rows or execute its SQL."""
    base_url = os.getenv("LLM_BASE_URL", "").rstrip("/")
    api_key = os.getenv("LLM_API_KEY", "")
    model = os.getenv("LLM_MODEL", "")
    safe_columns = [{"name": item.name, "kind": item.kind, "description": item.description, "distinct_count": item.distinct_count} for item in columns if item.kind != "id"]
    if not base_url or not api_key or not model or not safe_columns:
        return None, None
    prompt = {"role": "system", "content": "You are a read-only data intent parser. Return ONLY JSON: {dimension,metric,aggregation,chart_type,limit,clarification}. Select dimension and metric only from schema. Never return SQL. chart_type must be bar or line; aggregation must be sum, avg, or count. If ambiguous, set clarification."}
    user = {"role": "user", "content": json.dumps({"question": request.message, "prior_context": request.context[-1000:], "schema": safe_columns}, ensure_ascii=False)}
    payload = json.dumps({"model": model, "temperature": 0, "response_format": {"type": "json_object"}, "messages": [prompt, user]}).encode()
    try:
        http_request = urllib.request.Request(f"{base_url}/chat/completions", data=payload, headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}, method="POST")
        decoder = json.JSONDecoder()
        with urllib.request.urlopen(http_request, timeout=8) as response:
            # A compatible gateway may concatenate JSON response objects. Read only
            # the first complete response object; the intent itself is independently validated below.
            outer, _ = decoder.raw_decode(response.read().decode().lstrip())
            if not isinstance(outer, dict):
                return None, None
            content = outer["choices"][0]["message"]["content"]
        # Some compatible gateways append a second JSON fragment despite JSON mode.
        # Accept only the first complete object; validation below still rejects unsafe intent.
        parsed: object | None = None
        for offset, character in enumerate(content):
            if character != "{":
                continue
            try:
                candidate, _ = decoder.raw_decode(content[offset:])
                if isinstance(candidate, dict):
                    parsed = candidate
                    break
            except json.JSONDecodeError:
                continue
        if not isinstance(parsed, dict):
            return None, None
        if parsed.get("clarification"):
            return None, str(parsed["clarification"])[:400]
        aggregation = str(parsed.get("aggregation", "sum"))
        chart_type = str(parsed.get("chart_type", "bar"))
        if aggregation not in {"sum", "avg", "count"} or chart_type not in {"bar", "line"}:
            return None, None
        chart = ChartRequest(dimension=str(parsed["dimension"]), metric=str(parsed["metric"]), aggregation=cast(Literal["sum", "avg", "count"], aggregation), chart_type=cast(Literal["bar", "line"], chart_type), limit=min(30, max(1, int(parsed.get("limit", 12)))), filters=[])
        known = {item.name: item for item in columns}
        if chart.dimension not in known or chart.metric not in known or known[chart.dimension].kind not in {"time", "cat"} or known[chart.metric].kind != "num":
            return None, None
        if any(token in chart.metric.lower() for token in {"_id", "code", "key"}):
            return None, None
        if chart.chart_type == "line" and known[chart.dimension].kind != "time":
            chart.chart_type = "bar"
        return chart, None
    except (KeyError, ValueError, TypeError, urllib.error.URLError, urllib.error.HTTPError, TimeoutError):
        return None, None


def _selection_artifact(run_id: str) -> dict[str, object]:
    """Return schema-only user selections for this run; no selections are global."""
    path = RUN_STORE._dir(run_id) / "semantic-selection.json"
    if not path.exists():
        return {"selections": []}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data.get("selections"), list) else {"selections": []}
    except (OSError, json.JSONDecodeError):
        return {"selections": []}


def _named_columns(columns: list[ColumnProfile], message: str) -> set[str]:
    """Find explicit schema names, treating underscore/hyphen/space as equivalent."""
    normalized = re.sub(r"[^a-z0-9]+", " ", message.casefold()).strip()
    return {
        item.name for item in columns
        if re.search(rf"(?<![a-z0-9]){re.escape(re.sub(r'[^a-z0-9]+', ' ', item.name.casefold()).strip())}(?![a-z0-9])", normalized)
    }


def _semantic_clarification(columns: list[ColumnProfile], message: str, selections: list[dict[str, object]], language: Literal["en", "vi"] = "en") -> list[ClarificationOption]:
    """Ask only for roles not explicitly named in the schema or confirmed by this run's user."""
    normalized = message.lower()
    requested_measure = any(word in normalized for word in {"doanh thu", "revenue", "sales", "lợi nhuận", "profit", "margin", "chi phí", "cost", "số lượng", "quantity"})
    requested_dimension = any(word in normalized for word in {"theo", "by", "trend", "xu hướng", "tháng", "month", "ngày", "day"})
    named = _named_columns(columns, message)
    selected = {str(item.get("column")): str(item.get("role")) for item in selections}
    metric_named = any(item.kind == "num" and (item.name in named or selected.get(item.name) == "metric") for item in columns)
    dimension_named = any(item.kind in {"time", "cat"} and (item.name in named or selected.get(item.name) == "dimension") for item in columns)
    metrics = [item for item in columns if item.kind == "num" and not any(token in item.name.lower() for token in {"_id", "code", "key"})]
    dimensions = [item for item in columns if item.kind in {"time", "cat"}]
    catalog = business_semantic_catalog(columns)
    requested_business_metric = any(word in normalized for word in {"doanh thu", "revenue", "sales", "sale", "excl vat", "lợi nhuận", "profit", "margin", "chi phí", "cost", "số lượng", "quantity"})
    semantic_metric = (
        catalog.metric("sales") if any(word in normalized for word in {"doanh thu", "revenue", "sales", "sale", "excl vat"}) else
        catalog.metric("profit") if any(word in normalized for word in {"lợi nhuận", "profit", "margin"}) else
        catalog.metric("cost") if any(word in normalized for word in {"chi phí", "cost"}) else
        catalog.metric("quantity") if any(word in normalized for word in {"số lượng", "quantity"}) else None
    )
    if requested_business_metric and not metric_named and semantic_metric is None and metrics:
        return [ClarificationOption(column=item.name, label=item.name, reason="Không có alias chỉ tiêu kinh doanh được xác thực; hãy chọn một trường số." if language == "vi" else "No validated business alias matched; select the numeric field to use.", role="metric") for item in metrics[:4]]
    if not requested_measure and not metric_named and len(metrics) > 1:
        return [ClarificationOption(column=item.name, label=item.name, reason="Ứng viên chỉ tiêu số — hãy xác nhận ý nghĩa nghiệp vụ." if language == "vi" else "Numeric measure candidate — please confirm its business meaning.", role="metric") for item in metrics[:4]]
    if requested_measure and not requested_dimension and not dimension_named and len(dimensions) > 1:
        return [ClarificationOption(column=item.name, label=item.name, reason="Trường thời gian hoặc phân tách có thể phù hợp — hãy xác nhận." if language == "vi" else "Possible time or breakdown dimension — please confirm.", role="dimension") for item in dimensions[:4]]
    return []


def _top_stores_sales_by_region_request(columns: list[ColumnProfile], message: str, selections: list[dict[str, object]] | None = None) -> ChartRequest | None:
    """Resolve explicit ranked store/site + sales + optional region roles together.

    Returning None is deliberately not permission to substitute a date or first field:
    callers detect explicit-but-incomplete ranking language and ask for clarification.
    """
    normalized = canonical_field_name(message).replace("_", " ")
    top = re.search(r"(?:top|highest|cao nhat)\s+(\d{1,2})", normalized)
    wants_store = bool(re.search(r"\b(?:store|stores|site|sites|cua hang)\b", normalized))
    wants_sales = bool(re.search(r"\b(?:sales?|sale|revenue|doanh thu|excl vat|vat excluded)\b", normalized))
    wants_region = bool(re.search(r"\b(?:region|regions|vung|mien)\b", normalized))
    if not top or not wants_store or not wants_sales:
        return None
    catalog = business_semantic_catalog(columns)
    metric, store = catalog.metric("sales"), catalog.location()
    region = next((item for item in columns if item.kind == "cat" and canonical_field_name(item.name) in {"region", "sales_region", "area", "territory", "vung", "mien"}), None)
    selected_group = next((item for item in columns if item.kind == "cat" and any(selection.get("column") == item.name and selection.get("role") == "dimension" for selection in (selections or []))), None)
    grouping = region or selected_group
    if not metric or not store or (wants_region and not grouping):
        return None
    return ChartRequest(dimension=store.name, secondary_dimension=grouping.name if wants_region else None, metric=metric.name, aggregation="sum", chart_type="bar", limit=int(top.group(1)), limit_per_secondary=wants_region)


def _has_explicit_unresolved_ranking(message: str) -> bool:
    normalized = canonical_field_name(message).replace("_", " ")
    # This guard protects the compound store/region contract only. Other ranked
    # dimensions (for example channel by sales) proceed through normal resolution.
    return bool(re.search(r"(?:top|highest|cao nhat)\s+\d", normalized) and re.search(r"\b(?:store|stores|site|sites|cua hang|region|regions|vung|mien)\b", normalized))


def _ranking_clarification(columns: list[ColumnProfile], message: str, language: Literal["en", "vi"]) -> ChatResponse:
    """Name the missing requested role instead of issuing a generic rejection."""
    normalized = canonical_field_name(message).replace("_", " ")
    wants_region = bool(re.search(r"\b(?:region|regions|vung|mien)\b", normalized))
    catalog = business_semantic_catalog(columns)
    has_region = any(item.kind == "cat" and canonical_field_name(item.name) in {"region", "sales_region", "area", "territory", "vung", "mien"} for item in columns)
    candidates = [item for item in columns if item.kind == "cat" and item != catalog.location()][:4]
    options = [ClarificationOption(column=item.name, label=display_label(item.name), reason="Available categorical grouping candidate." if language == "en" else "Ứng viên nhóm phân loại hiện có.", role="dimension") for item in candidates]
    if wants_region and not has_region:
        names = ", ".join(display_label(item.name) for item in candidates) or ("no safe categorical fields" if language == "en" else "không có trường phân loại an toàn")
        answer = (f"Region is not available in this dataset, so I cannot rank stores within each Region. Available grouping columns: {names}." if language == "en" else f"Dataset này không có trường Region nên không thể xếp hạng cửa hàng theo từng Region. Các cột nhóm hiện có: {names}.")
        insight = "No aggregate ran; Region was not substituted with another field." if language == "en" else "Chưa chạy aggregate; Region không bị thay thế bằng trường khác."
        return ChatResponse(answer=answer, insight=insight, scope="Awaiting a grouping choice" if language == "en" else "Chờ chọn trường nhóm", caveats=[], clarification_options=options, mode="clarification")
    if language == "vi":
        return ChatResponse(answer="Chưa thể xác thực đủ cửa hàng, vùng và chỉ tiêu doanh số trong schema này. Hãy chọn các cột tương ứng.", insight="Chưa chạy aggregate để tránh thay thế yêu cầu xếp hạng bằng xu hướng ngày.", scope="Chờ xác nhận semantic", caveats=[], clarification_options=options, mode="clarification")
    return ChatResponse(answer="I could not verify every requested store, region, and sales role in this schema. Please select the matching columns.", insight="No aggregate ran, so the requested ranking is not replaced by a date trend.", scope="Awaiting semantic confirmation", caveats=[], clarification_options=options, mode="clarification")


def _validate_explicit_roles(message: str, chart: ChartRequest) -> bool:
    """Prevent LLM/deterministic plans from downgrading named ranking roles."""
    normalized = canonical_field_name(message).replace("_", " ")
    if re.search(r"(?:top|highest|cao nhat)\s+\d", normalized):
        if chart.chart_type != "bar" or chart.limit_per_secondary != bool(re.search(r"\b(?:region|regions|vung|mien)\b", normalized)):
            return False
    return True


def _chat_metric(columns: list[ColumnProfile], message: str, selections: list[dict[str, object]]) -> ColumnProfile | None:
    normalized = canonical_field_name(message).replace("_", " ")
    metrics = [item for item in columns if item.kind == "num" and not any(token in item.name.lower() for token in {"_id", "code", "key"})]
    selected = {str(item.get("column")) for item in selections if item.get("role") == "metric"}
    named = _named_columns(columns, message)
    direct = next((item for item in metrics if item.name in named or item.name in selected), None)
    if direct:
        return direct
    catalog = business_semantic_catalog(columns)
    intents = (("sales", ("doanh thu", "revenue", "sales", "sale", "excl vat")), ("profit", ("loi nhuan", "profit", "margin")), ("quantity", ("so luong", "quantity", "volume")), ("cost", ("chi phi", "cost")))
    for intent, words in intents:
        if any(word in normalized for word in words):
            return cast(ColumnProfile | None, catalog.metric(intent))
    return None


def _chat_dimension(columns: list[ColumnProfile], message: str, selections: list[dict[str, object]]) -> ColumnProfile | None:
    normalized = canonical_field_name(message).replace("_", " ")
    fields = [item for item in columns if item.kind in {"time", "cat"}]
    selected = {str(item.get("column")) for item in selections if item.get("role") == "dimension"}
    named = _named_columns(columns, message)
    direct = next((item for item in fields if item.name in named or item.name in selected), None)
    if direct:
        return direct
    if any(word in normalized for word in {"thang", "month", "ngay", "day", "trend", "xu huong", "6 thang", "nam"}):
        return cast(ColumnProfile | None, business_semantic_catalog(columns).time())
    for item in fields:
        words = canonical_field_name(item.name).split("_")
        if any(word in normalized for word in words if len(word) > 2):
            return item
    return fields[0] if len(fields) == 1 else None


def _output_intent(message: str) -> Literal["table", "chart"]:
    normalized = message.casefold()
    table_words = ("table", "tabular", "rows", "bảng", "dang bang", "dạng bảng")
    return "table" if any(word in normalized for word in table_words) else "chart"


def _starter_analysis_response(columns: list[ColumnProfile], message: str, language: Literal["en", "vi"] = "en") -> ChatResponse:
    """Return selectable, profile-only views without running a chart or calling an LLM."""
    proposals = analyst_proposals(columns, max_charts=5, language=language)
    english = language == "en"
    if english:
        answer = "Here are safe starter analysis views based on this dataset profile. Select a card to run its validated aggregate."
        insight = f"Prepared {len(proposals)} deterministic view{'s' if len(proposals) != 1 else ''}; no raw rows or chart query were sent to an LLM."
        scope = "Profile metadata only; no aggregate has run"
        caveat = "Cards exclude identifier and sensitive fields and require explicit approval before a chart runs."
    else:
        answer = "Đây là các góc nhìn phân tích khởi đầu an toàn từ profile dataset. Hãy chọn một thẻ để chạy aggregate đã được xác thực."
        insight = f"Đã chuẩn bị {len(proposals)} góc nhìn xác định; không gửi raw rows hoặc chart query tới LLM."
        scope = "Chỉ dùng metadata của profile; chưa chạy aggregate"
        caveat = "Các thẻ loại trừ trường định danh và nhạy cảm; chỉ chạy chart sau khi anh/chị phê duyệt."
    return ChatResponse(answer=answer, insight=insight, scope=scope, proposals=proposals, caveats=[caveat], mode="analysis", planner="deterministic")


def chart_insight(rows: list[dict[str, str | float | int]], chronological: bool, language: Literal["en", "vi"] = "en") -> tuple[str, list[str]]:
    if not rows:
        return ("Không có dữ liệu trong phạm vi đã chọn.", ["Không có aggregate nào phù hợp với filter và cột đã chọn."]) if language == "vi" else ("No data is available in the selected scope.", ["No aggregate matched the selected filters and columns."])
    values = [float(row["value"]) for row in rows]
    if chronological and len(rows) >= 2:
        first, last = values[0], values[-1]; delta = last - first; change = 0 if first == 0 else delta / abs(first) * 100; peak_index, trough_index = values.index(max(values)), values.index(min(values))
        if language == "vi":
            direction = "tăng" if delta >= 0 else "giảm"
            return (f"Giá trị cuối kỳ {direction} {percent(abs(change))} so với đầu kỳ.", [f"Từ {rows[0]['display_label']}: {compact_number(first)} đến {rows[-1]['display_label']}: {compact_number(last)} ({direction} {compact_number(abs(delta))}).", f"Đỉnh: {rows[peak_index]['display_label']} với {compact_number(values[peak_index])}; thấp nhất: {rows[trough_index]['display_label']} với {compact_number(values[trough_index])}."])
        direction = "increased" if delta >= 0 else "decreased"
        return (f"The final period {direction} by {percent(abs(change))} from the first period.", [f"From {rows[0]['display_label']}: {compact_number(first)} to {rows[-1]['display_label']}: {compact_number(last)} ({direction} by {compact_number(abs(delta))}).", f"Peak: {rows[peak_index]['display_label']} at {compact_number(values[peak_index])}; low: {rows[trough_index]['display_label']} at {compact_number(values[trough_index])}."])
    total = sum(values); top = rows[0]; share = 0 if total == 0 else float(top["value"]) / total * 100; second = values[1] if len(values) > 1 else None
    if language == "vi":
        bullets = [f"Dẫn đầu: {top['display_label']} với {compact_number(float(top['value']))}, chiếm {percent(share)} trong phần kết quả hiển thị."]
        if second is not None: bullets.append(f"Chênh lệch với hạng hai: {compact_number(float(top['value']) - second)}.")
        return f"{top['display_label']} đang dẫn đầu với {compact_number(float(top['value']))}.", bullets
    bullets = [f"Leader: {top['display_label']} at {compact_number(float(top['value']))}, representing {percent(share)} of the displayed total."]
    if second is not None: bullets.append(f"Gap to the runner-up: {compact_number(float(top['value']) - second)}.")
    return f"{top['display_label']} leads at {compact_number(float(top['value']))}.", bullets


@app.post("/api/runs/{run_id}/semantic-selection", response_model=ChatResponse)
def select_semantic_column(run_id: str, request: SemanticSelectionRequest) -> ChatResponse:
    """Persist a user choice inside one run, then resume only that run's pending question."""
    headers, rows = load_run(run_id)
    profile_data = profile_for_run(run_id, headers, rows)
    column = next((item for item in profile_data.columns if item.name == request.column), None)
    if column is None or is_sensitive_column(request.column):
        raise HTTPException(422, "Selected column is not available for semantic analysis.")
    if request.role == "metric" and column.kind != "num":
        raise HTTPException(422, "Selected metric must be a numeric schema column.")
    if request.role == "dimension" and column.kind not in {"time", "cat"}:
        raise HTTPException(422, "Selected dimension must be a time or categorical schema column.")
    state = _selection_artifact(run_id)
    selections = [item for item in cast(list[dict[str, object]], state["selections"]) if item.get("role") != request.role]
    selections.append({"column": column.name, "role": request.role, "provenance": "User"})
    continuation = state.get("continuation")
    pending_message = state.get("pending_message")
    if isinstance(continuation, dict):
        allowed_options = continuation.get("allowed_options")
        if not isinstance(allowed_options, list) or {"column": request.column, "role": request.role} not in allowed_options:
            raise HTTPException(422, "Selected option is not available for this question.")
        pending_message = continuation.get("message")
    if not isinstance(pending_message, str) or not pending_message:
        language = "vi" if request.language == "vi" else "en"
        return ChatResponse(
            answer="Phiên câu hỏi này đã kết thúc. Hãy gửi lại câu hỏi để tiếp tục." if language == "vi" else "This question session has ended. Please submit the question again to continue.",
            insight="Không có aggregate nào được chạy." if language == "vi" else "No aggregate has run.",
            scope="Cần gửi lại câu hỏi" if language == "vi" else "Question resubmission required",
            caveats=[],
            mode="clarification",
            planner="deterministic",
        )
    RUN_STORE.save_artifact_json(run_id, "semantic-selection.json", {"selections": selections})
    return chat_about_run(run_id, ChatRequest(message=pending_message, language=cast(Literal["en", "vi"], request.language)), allow_llm=False)


@app.post("/api/runs/{run_id}/chat", response_model=ChatResponse)
def chat_about_run(run_id: str, request: ChatRequest, *, allow_llm: bool = True) -> ChatResponse:
    """Constrained data conversation: natural language maps only to a validated aggregate."""
    headers, rows = load_run(run_id)
    profile_data = profile_for_run(run_id, headers, rows)
    if is_starter_analysis_request(request.message):
        return _starter_analysis_response(profile_data.columns, request.message, request.language)
    state = _selection_artifact(run_id)
    continuation = state.get("continuation")
    if isinstance(continuation, dict) and continuation.get("message") != request.message:
        # A new question must not inherit an old pending continuation, but retains
        # explicit run-scoped selections that the user has already confirmed.
        RUN_STORE.save_artifact_json(run_id, "semantic-selection.json", {"selections": state["selections"]})
        state = _selection_artifact(run_id)
    selections = cast(list[dict[str, object]], state["selections"])
    exact_request = _top_stores_sales_by_region_request(profile_data.columns, request.message, selections)
    if _has_explicit_unresolved_ranking(request.message) and exact_request is None:
        clarification = _ranking_clarification(profile_data.columns, request.message, request.language)
        RUN_STORE.save_artifact_json(run_id, "semantic-selection.json", {
            "version": 1,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "selections": selections,
            "continuation": {
                "message": request.message,
                "language": request.language,
                "requested_roles": ["store", "sales", "grouping"],
                "allowed_options": [{"column": option.column, "role": option.role} for option in clarification.clarification_options],
            },
        })
        return clarification
    clarification_options = [] if exact_request else _semantic_clarification(profile_data.columns, request.message, selections, request.language)
    if clarification_options:
        RUN_STORE.save_artifact_json(run_id, "semantic-selection.json", {"selections": selections, "pending_message": request.message})
        if request.language == "vi":
            return ChatResponse(answer="Có hơn một cột phù hợp nhưng chưa đủ chắc về ý nghĩa nghiệp vụ hoặc phạm vi. Hãy xác nhận cột muốn dùng.", insight="Chưa chạy aggregate để tránh tự suy diễn chỉ tiêu hoặc phạm vi.", scope="Chờ xác nhận semantic", caveats=["Lựa chọn chỉ áp dụng cho dataset/run hiện tại và được gắn provenance User."], clarification_options=clarification_options, mode="clarification")
        return ChatResponse(answer="More than one column could fit, but its business meaning or scope is not yet certain. Please confirm the column to use.", insight="No aggregate has run, to avoid inferring a metric or scope.", scope="Awaiting semantic confirmation", caveats=["Your selection applies only to this dataset/run and is recorded with User provenance."], clarification_options=clarification_options, mode="clarification")
    llm_request, clarification = _llm_chart_request(profile_data.columns, request) if allow_llm else (None, None)
    if clarification:
        return ChatResponse(answer=clarification, insight="AI cần xác nhận phạm vi trước khi chạy aggregate." if request.language == "vi" else "AI needs scope confirmation before running an aggregate.", scope="Chưa chạy phân tích" if request.language == "vi" else "No analysis has run", caveats=[], mode="clarification", planner="llm")
    planner = "llm" if llm_request else "deterministic"
    if exact_request:
        chart_request = exact_request
    elif llm_request:
        chart_request = llm_request
    else:
        metric = _chat_metric(profile_data.columns, request.message, selections)
        dimension = _chat_dimension(profile_data.columns, request.message, selections)
        if metric is None or dimension is None:
            return ChatResponse(answer="Cần một metric số và một trường thời gian hoặc dimension để phân tích. Có thể hỏi ‘Doanh thu theo tháng’ hoặc ‘Doanh thu theo kênh’." if request.language == "vi" else "I need a numeric metric and a time or categorical dimension to analyze. Try a question such as ‘Revenue by month’ or ‘Revenue by channel’.", insight="Chưa tìm được cặp metric/dimension an toàn trong dataset này." if request.language == "vi" else "No safe metric/dimension pair was found in this dataset.", scope="Chưa chạy phân tích" if request.language == "vi" else "No analysis has run", caveats=[], mode="clarification")
        chart_request = ChartRequest(dimension=dimension.name, metric=metric.name, aggregation="sum", chart_type="line" if dimension.kind == "time" else "bar", limit=12)
    output_intent = _output_intent(request.message)
    chart = build_chart(run_id, chart_request, request.language)
    metric = next(item for item in profile_data.columns if item.name == chart_request.metric)
    dimension = next(item for item in profile_data.columns if item.name == chart_request.dimension)
    scope = (f"{chart.aggregation.upper()} {metric.name} theo {dimension.name}; {chart.result_count} kết quả, xếp {chart.sort_mode}" if request.language == "vi" else f"{chart.aggregation.upper()} {metric.name} by {dimension.name}; {chart.result_count} results, sorted {chart.sort_mode}")
    insight = chart.insight_headline
    caveats = list(chart.warnings)
    if any(word in request.message.lower() for word in {"cùng kỳ", "year over year", "yoy", "năm ngoái"}):
        caveats.append("So sánh cùng kỳ cần một trường thời gian được chuẩn hoá theo tháng/năm; bản chat hiện trả xu hướng tổng hợp trước để bạn review phạm vi." if request.language == "vi" else "A year-over-year comparison needs a time field normalized by month/year; this response returns an aggregate trend for you to review the scope first.")
    answer = f"Đã chuẩn bị bảng {metric.name} theo {dimension.name}." if output_intent == "table" and request.language == "vi" else f"I prepared a table of {metric.name} by {dimension.name}." if output_intent == "table" else f"Đã phân tích {metric.name} theo {dimension.name}." if request.language == "vi" else f"I analyzed {metric.name} by {dimension.name}."
    return ChatResponse(answer=answer, insight=insight, scope=scope, title=chart.title, chart=None if output_intent == "table" else chart, table=chart.rows, caveats=caveats, mode="analysis", planner=cast(Literal["llm", "deterministic"], planner))


@app.post("/api/runs/{run_id}/chat/stream")
def stream_chat_about_run(run_id: str, request: ChatRequest) -> StreamingResponse:
    """Stream safe, user-meaningful progress states—not private model reasoning."""
    labels = {
        "en": {"received": "Question received", "planning": "Understanding the question against the dataset schema", "validating": "Checking safe columns and scope", "fallback": "Using the safe analysis planner for a fast response", "clarification": "A confirmation is needed before analysis", "aggregating": "Aggregating data on the server", "insights": "Forming insight from verified aggregates", "completed": "Analysis complete"},
        "vi": {"received": "Đã nhận câu hỏi", "planning": "Đang hiểu câu hỏi theo schema dataset", "validating": "Đang kiểm tra cột và phạm vi an toàn", "fallback": "Đang dùng bộ phân tích an toàn để phản hồi nhanh", "clarification": "Cần xác nhận thêm trước khi chạy phân tích", "aggregating": "Đã tổng hợp dữ liệu trên server", "insights": "Đang tạo insight từ aggregate đã xác thực", "completed": "Hoàn tất phân tích"},
    }[request.language]
    def event(name: str, payload: dict[str, object]) -> str:
        return f"event: {name}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

    def generate():
        yield event("received", {"label": labels["received"]})
        yield event("planning", {"label": labels["planning"]})
        yield event("validating", {"label": labels["validating"]})
        try:
            # The synchronous LLM client cannot yield while waiting. Stream stays responsive
            # by using the deterministic, schema-validated planner for this transport.
            yield event("fallback_planning", {"label": labels["fallback"]})
            response = chat_about_run(run_id, request, allow_llm=True)
            if response.mode == "clarification":
                yield event("clarification", {"label": labels["clarification"], "response": response.model_dump()})
            else:
                yield event("aggregating", {"label": labels["aggregating"]})
                yield event("insights", {"label": labels["insights"]})
                yield event("completed", {"label": labels["completed"], "response": response.model_dump()})
        except HTTPException as error:
            yield event("error", {"label": str(error.detail) if error.status_code < 500 else "Không thể hoàn tất phân tích lúc này."})
        except Exception:
            yield event("error", {"label": "Không thể hoàn tất phân tích lúc này."})
    return StreamingResponse(generate(), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


if STATIC_DIR.is_dir():
    app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="web")
